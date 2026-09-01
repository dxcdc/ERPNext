"""Relatórios operacionais CDC construídos sobre o escopo nativo de armazéns."""

import csv
import io
import json
from collections import defaultdict
from html import escape

import frappe
from frappe.utils import flt, getdate, now_datetime

from cdc_theme.api import (
    CDC_PROJECTS,
    _permitted_leaf_warehouses,
    _require_read_permission,
    _require_stock_reports_access,
    _warehouse_project,
)


REPORT_MAX_ROWS = 50000
PDF_DETAIL_LIMIT = 250
MOVEMENT_TYPES = {"all", "receipt", "issue", "transfer"}
SCOPE_MODES = {"warehouses", "group", "project"}


def _parse_list(value):
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except (TypeError, ValueError):
            value = [part.strip() for part in value.split(",") if part.strip()]
    if not isinstance(value, (list, tuple, set)):
        return []
    return list(dict.fromkeys(str(item).strip() for item in value if str(item).strip()))


def _report_permitted_leaf_warehouses():
    """Mantém Stock User fechado aos Warehouse vinculados por User Permission."""
    permitted = _permitted_leaf_warehouses()
    roles = set(frappe.get_roles(frappe.session.user))
    managers = {"System Manager", "Stock Manager", "Gestor de Estoque"}
    if "Stock User" not in roles or roles.intersection(managers):
        return permitted

    linked = set(frappe.get_all(
        "User Permission",
        filters={
            "user": frappe.session.user,
            "allow": "Warehouse",
        },
        pluck="for_value",
    ))
    if not linked:
        return set()

    warehouses = frappe.get_all(
        "Warehouse",
        fields=["name", "is_group", "lft", "rgt"],
        filters={"name": ["in", list(linked)]},
    )
    linked_leaves = set()
    for warehouse in warehouses:
        if not int(warehouse.is_group or 0):
            linked_leaves.add(warehouse.name)
            continue
        descendants = frappe.get_all(
            "Warehouse",
            filters={
                "is_group": 0,
                "lft": [">", warehouse.lft],
                "rgt": ["<", warehouse.rgt],
            },
            pluck="name",
        )
        linked_leaves.update(descendants)
    return permitted.intersection(linked_leaves)


def _warehouse_context():
    permitted = _report_permitted_leaf_warehouses()
    rows = frappe.get_list(
        "Warehouse",
        fields=[
            "name", "warehouse_name", "company", "disabled", "is_group",
            "parent_warehouse", "lft", "rgt",
        ],
        order_by="name asc",
        limit_page_length=0,
    )
    leaves = [row for row in rows if not int(row.is_group or 0) and row.name in permitted]
    groups = []
    for group in (row for row in rows if int(row.is_group or 0)):
        descendants = [
            row.name for row in leaves
            if group.lft is not None and group.rgt is not None
            and row.lft is not None and group.lft < row.lft < group.rgt
        ]
        if descendants:
            groups.append({
                "name": group.name,
                "label": (group.warehouse_name or group.name).removesuffix(" - C"),
                "warehouses": descendants,
            })
    return leaves, groups, permitted


def _resolve_scope(scope_mode, warehouses=None, group=None, project=None):
    mode = str(scope_mode or "warehouses").strip().lower()
    if mode not in SCOPE_MODES:
        frappe.throw("Modo de seleção inválido.", frappe.ValidationError)
    leaves, groups, permitted = _warehouse_context()
    leaf_names = {row.name for row in leaves}
    if mode == "warehouses":
        selected = set(_parse_list(warehouses))
        if not selected:
            frappe.throw("Selecione ao menos um armazém.", frappe.ValidationError)
    elif mode == "group":
        selected_group = next((item for item in groups if item["name"] == str(group or "").strip()), None)
        if not selected_group:
            frappe.throw("Grupo indisponível para o usuário atual.", frappe.PermissionError)
        selected = set(selected_group["warehouses"])
    else:
        selected_project = str(project or "").strip()
        available_projects = {_warehouse_project(row.name) for row in leaves}
        if selected_project not in available_projects:
            frappe.throw("Projeto indisponível para o usuário atual.", frappe.PermissionError)
        selected = {row.name for row in leaves if _warehouse_project(row.name) == selected_project}
    forbidden = selected - permitted
    if forbidden or not selected.issubset(leaf_names):
        frappe.throw("Um ou mais armazéns estão indisponíveis para o usuário atual.", frappe.PermissionError)
    return sorted(selected), leaves, groups


def _validate_period(from_date, to_date):
    try:
        start = getdate(from_date)
        end = getdate(to_date)
    except Exception:
        frappe.throw("Informe um período válido.", frappe.ValidationError)
    if not from_date or not to_date:
        frappe.throw("Informe as datas inicial e final.", frappe.ValidationError)
    if start > end:
        frappe.throw("A data inicial não pode ser posterior à data final.", frappe.ValidationError)
    return start, end


def _report_query(selected, start, end):
    placeholders = ", ".join(["%s"] * len(selected))
    meta = frappe.get_meta("Stock Entry")
    order_field = "se.idpedido_ongsys" if meta.has_field("idpedido_ongsys") else "NULL"
    title_field = "se.titulo_ongsys" if meta.has_field("titulo_ongsys") else "NULL"
    values = [str(start), str(end), *selected, *selected, REPORT_MAX_ROWS + 1]
    return frappe.db.sql(
        f"""
        SELECT
            se.name AS stock_entry,
            se.posting_date,
            se.posting_time,
            se.purpose,
            se.stock_entry_type,
            se.owner,
            {order_field} AS ongsys_order,
            {title_field} AS ongsys_title,
            sed.idx,
            sed.item_code,
            sed.item_name,
            sed.uom,
            sed.qty,
            sed.s_warehouse,
            sed.t_warehouse
        FROM `tabStock Entry` se
        INNER JOIN `tabStock Entry Detail` sed ON sed.parent = se.name
        WHERE se.docstatus = 1
          AND se.posting_date BETWEEN %s AND %s
          AND (
            sed.s_warehouse IN ({placeholders})
            OR sed.t_warehouse IN ({placeholders})
          )
        ORDER BY se.posting_date ASC, se.posting_time ASC, se.name ASC, sed.idx ASC
        LIMIT %s
        """,
        values=values,
        as_dict=True,
    )


def _movement_rows(raw_rows, selected, permitted, movement_type):
    requested = str(movement_type or "all").strip().lower()
    if requested not in MOVEMENT_TYPES:
        frappe.throw("Tipo de movimentação inválido.", frappe.ValidationError)
    selected_set = set(selected)
    result = []

    def public_warehouse(value):
        return value if value in permitted else ""

    def add(row, warehouse, category, direction):
        if requested not in {"all", category}:
            return
        result.append({
            "stock_entry": row.stock_entry,
            "posting_date": str(row.posting_date or ""),
            "posting_time": str(row.posting_time or ""),
            "category": category,
            "direction": direction,
            "warehouse": warehouse,
            "source_warehouse": public_warehouse(row.s_warehouse),
            "target_warehouse": public_warehouse(row.t_warehouse),
            "item_code": row.item_code or "",
            "item_name": row.item_name or "",
            "uom": row.uom or "",
            "quantity": flt(row.qty, 6),
            "ongsys_order": str(row.ongsys_order or ""),
            "ongsys_title": row.ongsys_title or "",
            "owner": row.owner or "",
        })

    for row in raw_rows:
        source_selected = row.s_warehouse in selected_set
        target_selected = row.t_warehouse in selected_set
        is_transfer = bool(row.s_warehouse and row.t_warehouse)
        if is_transfer:
            if source_selected:
                add(row, row.s_warehouse, "transfer", "Saída por transferência")
            if target_selected:
                add(row, row.t_warehouse, "transfer", "Entrada por transferência")
        elif target_selected:
            add(row, row.t_warehouse, "receipt", "Entrada")
        elif source_selected:
            add(row, row.s_warehouse, "issue", "Saída")
    return result


def _summary(rows, selected, start, end):
    by_warehouse = defaultdict(lambda: {"entries": 0.0, "issues": 0.0, "transfers_in": 0.0, "transfers_out": 0.0})
    for row in rows:
        bucket = by_warehouse[row["warehouse"]]
        if row["category"] == "receipt":
            bucket["entries"] += row["quantity"]
        elif row["category"] == "issue":
            bucket["issues"] += row["quantity"]
        elif row["direction"].startswith("Entrada"):
            bucket["transfers_in"] += row["quantity"]
        else:
            bucket["transfers_out"] += row["quantity"]
    return {
        "from_date": str(start),
        "to_date": str(end),
        "warehouses": selected,
        "warehouse_count": len(selected),
        "documents": len({row["stock_entry"] for row in rows}),
        "movement_lines": len(rows),
        "distinct_items": len({row["item_code"] for row in rows if row["item_code"]}),
        "receipt_quantity": flt(sum(row["quantity"] for row in rows if row["category"] == "receipt"), 6),
        "issue_quantity": flt(sum(row["quantity"] for row in rows if row["category"] == "issue"), 6),
        "transfer_documents": len({row["stock_entry"] for row in rows if row["category"] == "transfer"}),
        "by_warehouse": [{"warehouse": name, **values} for name, values in sorted(by_warehouse.items())],
        "generated_at": str(now_datetime()),
    }


def _build_report(scope_mode, warehouses, group, project, from_date, to_date, movement_type):
    _require_stock_reports_access()
    _require_read_permission("Warehouse")
    _require_read_permission("Stock Entry")
    start, end = _validate_period(from_date, to_date)
    selected, _leaves, _groups = _resolve_scope(scope_mode, warehouses, group, project)
    permitted = _report_permitted_leaf_warehouses()
    raw_rows = _report_query(selected, start, end)
    if len(raw_rows) > REPORT_MAX_ROWS:
        frappe.throw(
            f"O período ultrapassou {REPORT_MAX_ROWS} linhas. Reduza o período ou o número de armazéns.",
            frappe.ValidationError,
        )
    rows = _movement_rows(raw_rows, selected, permitted, movement_type)
    return {"summary": _summary(rows, selected, start, end), "rows": rows}


@frappe.whitelist()
def get_stock_movement_report_options():
    _require_stock_reports_access()
    _require_read_permission("Warehouse")
    _require_read_permission("Stock Entry")
    leaves, groups, _permitted = _warehouse_context()
    projects = [project for project in CDC_PROJECTS if any(_warehouse_project(row.name) == project for row in leaves)]
    return {
        "warehouses": [{
            "name": row.name,
            "label": (row.warehouse_name or row.name).removesuffix(" - C"),
            "company": row.company or "",
            "disabled": int(row.disabled or 0),
        } for row in leaves],
        "groups": groups,
        "projects": projects,
        "formats": ["pdf", "xlsx", "csv"],
        "movement_types": ["all", "receipt", "issue", "transfer"],
    }


@frappe.whitelist()
def preview_stock_movement_report(
    scope_mode="warehouses", warehouses=None, group=None, project=None,
    from_date=None, to_date=None, movement_type="all",
):
    report = _build_report(scope_mode, warehouses, group, project, from_date, to_date, movement_type)
    return {"summary": report["summary"], "rows": report["rows"][:100], "preview_limit": 100}


def _csv_content(report):
    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow([
        "Data", "Lançamento", "Movimentação", "Armazém", "Origem", "Destino",
        "Código do item", "Nome do item", "Unidade", "Quantidade", "Pedido ONGSYS", "Responsável",
    ])
    for row in report["rows"]:
        writer.writerow([
            row["posting_date"], row["stock_entry"], row["direction"], row["warehouse"],
            row["source_warehouse"], row["target_warehouse"], row["item_code"], row["item_name"],
            row["uom"], row["quantity"], row["ongsys_order"], row["owner"],
        ])
    return output.getvalue().encode("utf-8")


def _xlsx_content(report):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resumo"
    summary = report["summary"]
    summary_sheet.append(["Levantamento de movimentações de estoque"])
    summary_sheet["A1"].font = Font(bold=True, size=14)
    for label, value in (
        ("Data inicial", summary["from_date"]), ("Data final", summary["to_date"]),
        ("Armazéns", summary["warehouse_count"]), ("Lançamentos", summary["documents"]),
        ("Linhas de movimentação", summary["movement_lines"]), ("Itens distintos", summary["distinct_items"]),
        ("Quantidade de entradas", summary["receipt_quantity"]), ("Quantidade de saídas", summary["issue_quantity"]),
        ("Transferências", summary["transfer_documents"]), ("Gerado em", summary["generated_at"]),
    ):
        summary_sheet.append([label, value])
    summary_sheet.append([])
    summary_sheet.append(["Armazém", "Entradas", "Saídas", "Transferências recebidas", "Transferências enviadas"])
    for item in summary["by_warehouse"]:
        summary_sheet.append([item["warehouse"], item["entries"], item["issues"], item["transfers_in"], item["transfers_out"]])

    movement_sheet = workbook.create_sheet("Movimentações")
    headers = [
        "Data", "Lançamento", "Movimentação", "Armazém", "Origem", "Destino",
        "Código do item", "Nome do item", "Unidade", "Quantidade", "Pedido ONGSYS", "Responsável",
    ]
    movement_sheet.append(headers)
    for row in report["rows"]:
        movement_sheet.append([
            row["posting_date"], row["stock_entry"], row["direction"], row["warehouse"],
            row["source_warehouse"], row["target_warehouse"], row["item_code"], row["item_name"],
            row["uom"], row["quantity"], row["ongsys_order"], row["owner"],
        ])

    item_totals = defaultdict(lambda: {"item_name": "", "uom": "", "quantity": 0.0})
    for row in report["rows"]:
        key = (row["warehouse"], row["item_code"], row["direction"])
        item_totals[key]["item_name"] = row["item_name"]
        item_totals[key]["uom"] = row["uom"]
        item_totals[key]["quantity"] += row["quantity"]
    item_sheet = workbook.create_sheet("Itens")
    item_sheet.append(["Armazém", "Código", "Item", "Movimentação", "Unidade", "Quantidade"])
    for (warehouse, item_code, direction), values in sorted(item_totals.items()):
        item_sheet.append([
            warehouse, item_code, values["item_name"], direction,
            values["uom"], flt(values["quantity"], 6),
        ])

    for sheet in (summary_sheet, movement_sheet, item_sheet):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 48)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_content(report):
    from frappe.utils.pdf import get_pdf

    summary = report["summary"]
    warehouse_rows = "".join(
        f"<tr><td>{escape(item['warehouse'])}</td><td>{item['entries']}</td><td>{item['issues']}</td>"
        f"<td>{item['transfers_in']}</td><td>{item['transfers_out']}</td></tr>"
        for item in summary["by_warehouse"]
    ) or "<tr><td colspan='5'>Nenhuma movimentação encontrada.</td></tr>"
    detail_rows = "".join(
        f"<tr><td>{escape(row['posting_date'])}</td><td>{escape(row['stock_entry'])}</td>"
        f"<td>{escape(row['direction'])}</td><td>{escape(row['warehouse'])}</td>"
        f"<td>{escape(row['item_code'])}</td><td>{escape(row['item_name'])}</td>"
        f"<td>{row['quantity']}</td><td>{escape(row['uom'])}</td></tr>"
        for row in report["rows"][:PDF_DETAIL_LIMIT]
    ) or "<tr><td colspan='8'>Nenhuma movimentação encontrada.</td></tr>"
    truncated = len(report["rows"]) > PDF_DETAIL_LIMIT
    html = f"""
    <html><head><meta charset="utf-8"><style>
    body {{ font-family: sans-serif; color: #172033; font-size: 9pt; }}
    h1 {{ font-size: 18pt; margin-bottom: 4px; }} h2 {{ font-size: 12pt; margin-top: 18px; }}
    .meta {{ color: #526077; margin-bottom: 14px; }}
    .cards {{ width: 100%; margin: 10px 0; }} .cards td {{ padding: 7px; border: 1px solid #d8e0ec; }}
    table.data {{ width: 100%; border-collapse: collapse; }}
    table.data th {{ background: #2563eb; color: white; padding: 5px; text-align: left; }}
    table.data td {{ border: 1px solid #d8e0ec; padding: 4px; }}
    .note {{ color: #6b7280; font-size: 8pt; margin-top: 8px; }}
    </style></head><body>
    <h1>Movimentações de Estoque</h1>
    <div class="meta">Período: {summary['from_date']} a {summary['to_date']} · Gerado em {escape(summary['generated_at'])}</div>
    <table class="cards"><tr><td><b>{summary['warehouse_count']}</b><br>Armazéns</td>
    <td><b>{summary['documents']}</b><br>Lançamentos</td><td><b>{summary['movement_lines']}</b><br>Movimentações</td>
    <td><b>{summary['distinct_items']}</b><br>Itens distintos</td></tr></table>
    <h2>Resumo por armazém</h2><table class="data"><thead><tr><th>Armazém</th><th>Entradas</th><th>Saídas</th><th>Transf. recebidas</th><th>Transf. enviadas</th></tr></thead><tbody>{warehouse_rows}</tbody></table>
    <h2>Detalhamento</h2><table class="data"><thead><tr><th>Data</th><th>Lançamento</th><th>Tipo</th><th>Armazém</th><th>Código</th><th>Item</th><th>Qtd.</th><th>Un.</th></tr></thead><tbody>{detail_rows}</tbody></table>
    {"<p class='note'>O PDF exibe as primeiras 250 linhas. Utilize XLSX ou CSV para o detalhamento completo.</p>" if truncated else ""}
    </body></html>
    """
    return get_pdf(html, options={"orientation": "Landscape", "page-size": "A4", "footer-right": "Página [page] de [toPage]"})


@frappe.whitelist()
def download_stock_movement_report(
    file_format="xlsx", scope_mode="warehouses", warehouses=None, group=None, project=None,
    from_date=None, to_date=None, movement_type="all",
):
    requested_format = str(file_format or "xlsx").strip().lower()
    if requested_format not in {"pdf", "xlsx", "csv"}:
        frappe.throw("Formato de arquivo inválido.", frappe.ValidationError)
    report = _build_report(scope_mode, warehouses, group, project, from_date, to_date, movement_type)
    generators = {"pdf": _pdf_content, "xlsx": _xlsx_content, "csv": _csv_content}
    content = generators[requested_format](report)
    filename = f"movimentacoes-estoque-{report['summary']['from_date']}-{report['summary']['to_date']}.{requested_format}"
    frappe.local.response.filename = filename
    frappe.local.response.filecontent = content
    frappe.local.response.type = "download"
    frappe.local.response.display_content_as = "attachment"
