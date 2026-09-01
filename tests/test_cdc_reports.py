import importlib.util
import io
import sys
import types
import unittest
from datetime import date, datetime
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
MODULE_PATH = ROOT / "apps/cdc_theme/cdc_theme/reports.py"


def load_reports():
    frappe = types.ModuleType("frappe")
    frappe.whitelist = lambda *args, **kwargs: (lambda function: function) if not args else args[0]
    frappe.ValidationError = type("ValidationError", (Exception,), {})
    frappe.PermissionError = type("PermissionError", (Exception,), {})
    frappe.throw = lambda message, error=None: (_ for _ in ()).throw((error or Exception)(message))
    frappe.session = types.SimpleNamespace(user="stock.user@example.com")
    frappe.get_roles = lambda user=None: ["System Manager"]
    frappe.get_all = lambda *args, **kwargs: []
    frappe.utils = types.ModuleType("frappe.utils")
    frappe.utils.flt = lambda value, precision=6: round(float(value or 0), precision)
    frappe.utils.getdate = lambda value=None: value if isinstance(value, date) else date.fromisoformat(str(value))
    frappe.utils.now_datetime = lambda: datetime(2026, 8, 31, 16, 0, 0)

    api = types.ModuleType("cdc_theme.api")
    api.CDC_PROJECTS = ("Projeto Cais", "Institucional / Geral")
    api._permitted_leaf_warehouses = lambda: {"A - C", "B - C"}
    api._require_read_permission = lambda doctype: None
    api._require_stock_reports_access = lambda: None
    api._warehouse_project = lambda warehouse: "Projeto Cais" if "CAIS" in warehouse else "Institucional / Geral"
    package = types.ModuleType("cdc_theme")
    package.__path__ = []

    previous = {name: sys.modules.get(name) for name in ("frappe", "frappe.utils", "cdc_theme", "cdc_theme.api")}
    sys.modules.update({"frappe": frappe, "frappe.utils": frappe.utils, "cdc_theme": package, "cdc_theme.api": api})
    try:
        spec = importlib.util.spec_from_file_location("cdc_reports_test", MODULE_PATH)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module
    finally:
        for name, value in previous.items():
            if value is None:
                sys.modules.pop(name, None)
            else:
                sys.modules[name] = value


def movement(**values):
    defaults = {
        "stock_entry": "MAT-STE-0001", "posting_date": date(2026, 8, 1),
        "posting_time": "10:00:00", "purpose": "", "stock_entry_type": "",
        "owner": "user@example.com", "ongsys_order": "", "ongsys_title": "",
        "idx": 1, "item_code": "ITEM-1", "item_name": "Item teste", "uom": "Un",
        "qty": 1, "s_warehouse": None, "t_warehouse": None,
    }
    defaults.update(values)
    return types.SimpleNamespace(**defaults)


class CDCReportsTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.reports = load_reports()

    def test_entries_issues_and_transfer_sides_share_one_contract(self):
        rows = [
            movement(stock_entry="REC-1", qty=5, t_warehouse="A - C"),
            movement(stock_entry="ISS-1", qty=3, s_warehouse="A - C"),
            movement(stock_entry="TRA-1", qty=2, s_warehouse="A - C", t_warehouse="B - C"),
        ]
        result = self.reports._movement_rows(rows, ["A - C", "B - C"], {"A - C", "B - C"}, "all")
        self.assertEqual(4, len(result))
        self.assertEqual(["receipt", "issue", "transfer", "transfer"], [row["category"] for row in result])
        summary = self.reports._summary(result, ["A - C", "B - C"], date(2026, 8, 1), date(2026, 8, 31))
        self.assertEqual(3, summary["documents"])
        self.assertEqual(4, summary["movement_lines"])
        self.assertEqual(5, summary["receipt_quantity"])
        self.assertEqual(3, summary["issue_quantity"])
        self.assertEqual(1, summary["transfer_documents"])
        by_warehouse = {row["warehouse"]: row for row in summary["by_warehouse"]}
        self.assertEqual(2, by_warehouse["A - C"]["transfers_out"])
        self.assertEqual(2, by_warehouse["B - C"]["transfers_in"])

    def test_transfer_filter_and_unauthorized_endpoint_masking(self):
        rows = [movement(stock_entry="TRA-2", qty=4, s_warehouse="FORBIDDEN - C", t_warehouse="A - C")]
        result = self.reports._movement_rows(rows, ["A - C"], {"A - C"}, "transfer")
        self.assertEqual(1, len(result))
        self.assertEqual("", result[0]["source_warehouse"])
        self.assertEqual("A - C", result[0]["target_warehouse"])

    def test_stock_user_scope_requires_linked_warehouse(self):
        original_roles = self.reports.frappe.get_roles
        original_get_all = self.reports.frappe.get_all
        original_permitted = self.reports._permitted_leaf_warehouses
        try:
            self.reports.frappe.get_roles = lambda user=None: ["Stock User"]
            self.reports._permitted_leaf_warehouses = lambda: {"A - C", "B - C"}

            def get_all(doctype, **kwargs):
                if doctype == "User Permission":
                    return ["A - C"]
                return [types.SimpleNamespace(name="A - C", is_group=0, lft=2, rgt=3)]

            self.reports.frappe.get_all = get_all
            self.assertEqual({"A - C"}, self.reports._report_permitted_leaf_warehouses())

            self.reports.frappe.get_all = lambda *args, **kwargs: []
            self.assertEqual(set(), self.reports._report_permitted_leaf_warehouses())
        finally:
            self.reports.frappe.get_roles = original_roles
            self.reports.frappe.get_all = original_get_all
            self.reports._permitted_leaf_warehouses = original_permitted

    def test_csv_and_xlsx_use_the_same_rows(self):
        rows = self.reports._movement_rows(
            [movement(stock_entry="REC-Ç", qty=1.5, item_name="Álcool", t_warehouse="A - C")],
            ["A - C"], {"A - C"}, "all",
        )
        report = {"rows": rows, "summary": self.reports._summary(rows, ["A - C"], date(2026, 8, 1), date(2026, 8, 31))}
        csv_content = self.reports._csv_content(report)
        self.assertTrue(csv_content.startswith(b"\xef\xbb\xbf"))
        self.assertIn("Álcool".encode(), csv_content)
        xlsx_content = self.reports._xlsx_content(report)
        self.assertTrue(xlsx_content.startswith(b"PK"))
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(xlsx_content), read_only=True)
        self.assertEqual(["Resumo", "Movimentações", "Itens"], workbook.sheetnames)
        self.assertEqual("Álcool", workbook["Movimentações"]["H2"].value)


if __name__ == "__main__":
    unittest.main()
